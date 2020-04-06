from urllib.request import urlopen, pathname2url
from bs4 import BeautifulSoup
from openpyxl import Workbook
publicationExcel = Workbook()
newsExcel = Workbook()
odsExcel = Workbook()
titulo = ""
descricao = ""
ano = 2017
status = 0
ods = []
noticias = []
def buildOds(ods1, line, pub_id):
    publicacao_ods = odsExcel.active
    publicacao_ods.title = "Publicacao_ods"
    publicacao_ods['A1'] = 'publicacao_id'
    publicacao_ods['B1'] = 'ods'
    publicacao_ods['A'+str(line)] = str(pub_id)
    aux = ''
    i = 0
    for i in range(len(ods1)):
        aux = str(aux) + ods1[i]
    publicacao_ods['B'+str(line)] = aux
    line+=1
    odsExcel.save("arquvo1.xlsx")
    return "Ods ok"

def buildNews(noticias,line, pub_id):
    publicacao_noticias = newsExcel.active
    publicacao_noticias.title = "Publicacao_noticias"
    publicacao_noticias['A1'] = 'publicacao_id'
    publicacao_noticias['B1'] = 'noticias'
    publicacao_noticias['A'+str(line)] = pub_id
    aux = ''
    i = 0
    for i in range(len(noticias)):
        aux = aux + str(noticias[i])
    aux1 = aux.replace('[','').replace(']','').replace("'", "")
    publicacao_noticias['B'+str(line)] = aux1
    newsExcel.save("arquivo2.xlsx")
    return "Notícias concluídas..."

def buildPublication(titulo, descricao, status,line):
    planilhaPublicacao = publicationExcel.active
    planilhaPublicacao.title="Publicacao"
    planilhaPublicacao['A1'] = 'titulo'
    planilhaPublicacao['B1'] = 'descricao'
    planilhaPublicacao['C1'] = 'ano'
    planilhaPublicacao['D1'] = 'status'
    planilhaPublicacao['E1'] = 'responsavel_usuario'
    planilhaPublicacao['F1'] = 'secretaria_codSecretaria'
    planilhaPublicacao['A'+str(line)] = str(titulo.replace('[','').replace(']',''))
    planilhaPublicacao['B' + str(line)] = descricao.replace('[','').replace(']','')
    planilhaPublicacao['C' + str(line)] = 2017
    planilhaPublicacao['D' + str(line)] = status
    planilhaPublicacao['E' + str(line)] = 'Admin'
    planilhaPublicacao['F' + str(line)] = '1'
    publicationExcel.save("arquivo3.xlsx")
    return "Espero que tenha funcionado!"

def main():
    #variáveis
    i = 0
    j = 0
    line = 2
    pub_id = 1
    # Pegando HTML
    '''html = urlopen("{urlPage")'''
    html = urlopen("{urlPage}")
    htmlPage = BeautifulSoup(html.read(), "html5lib")
    findAllPublication = htmlPage.findAll("div", {"class": "row vertical-content"})  # Encontra todas as publicações
    publicationLen = (len(findAllPublication))  # pega a quantidade de publicação

    while i < publicationLen:
        publication = findAllPublication[i]
        publicationTitle = publication.findAll("h3")  # pega o título da publicação
        publicationDescription = publication.findAll("p", {"class": "text-muted mb-4 features-desc"});  # pega a descrição da publicação
        #publicationStatus = publication.findAll("div", {"class": "progress-bar progress-bar-striped progress-bar-animated"})  # pega a porcentagem concluída
        publicationOds = publication.findAll("img", {"class": "imgODS"})  # pega um vetor de ODS
        #publicationNews = publication.findAll("a", {"target": "_blank"})
        #---------------- preenchendo para excel --------------#
        titulo = str(publicationTitle).replace('<h3 class="mb-4 font-weight-light" style="text-align: justify">','').replace('</h3>', '')
        descricao = str(publicationDescription).replace('<p class="text-muted mb-4 features-desc" style="text-align: justify">','').replace(' </p>','')
        status = publication.find("div",class_="progress-bar progress-bar-striped progress-bar-animated")["aria-valuenow"]
        publicationNews = publication.find_all("a", href=True)
        p = [a['href'] for a in publication.find_all('a', href=True) if a.text]
        if publicationNews:
            noticias.append(p)
        else:
            noticias.append('#')
        odsImg = str(publicationOds).replace('.png"="" assetsland="" class="imgODS" images="" ods="" src="','').replace('                  &lt;c:url value="/>','').replace('<img ','').replace(' ','')
        ods1 = odsImg.replace('[','').replace(']','').split()
        buildPublication(titulo, descricao, status, line)
        buildNews(noticias, line, pub_id)
        buildOds(ods1, line, pub_id)
        line+=1
        pub_id+=1
        i+=1
        noticias.clear()
    print("Concluído")
main()
